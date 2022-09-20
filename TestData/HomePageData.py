import openpyxl
#With static methods we can access methods with class name no need to create object of the class

class HomePageData:

    test_HomePage_data = [{"firstname":"neethi","email":"prince","gender":"Female"}, {"firstname":"Prince", "email":"Mohandas", "gender":"Male"}]
    @staticmethod
    def getTestData(test_case_name):#self parameter is not required because it is a static method only for non-static methods self parameter is required
        dict={}
        book = openpyxl.load_workbook("C:\\Users\\Techment Technology\\OneDrive\\Desktop\\techment project\\API\\ExcelDemo.xlsx")  # workbook is collection of sheets
        sheet=book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "Testcase4":
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i,column=j).value  # from excel values trying to build dictionary
                    # print(sheet.cell(row=i,column=j).value)#printing all the values from sheet

        return[dict]
'''
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Owner\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
'''
