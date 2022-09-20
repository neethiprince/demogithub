import openpyxl
book=openpyxl.load_workbook("C:\\Users\\Techment Technology\\OneDrive\\Desktop\\techment project\\API\\ExcelDemo.xlsx")#workbook is collection of sheets
sheet=book.active#giving control to sheet object because there might be many sheets out of which one will be active
dict={}
cell=sheet.cell(row=1,column=2)#control is shifted to this cell object
print(cell.value)
cell2=sheet.cell(row=2,column=2).value="Neethi"
cell3=sheet.cell(row=2,column=3).value="Prince"
cell4=sheet.cell(row=2,column=4).value="neethi@123"
print(cell2)
print(cell3)
print(cell4)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A8'].value)#instead of sheet.cell(row=2,column=2) we can use sheet['A8']
'''
for i in range(1,sheet.max_row+1):
    print(sheet.cell(row=i,column=1).value)#iteration for only row value and column 1 is fixed
'''
#from excel values we are creating dictionary
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value =="Testcase2":
        for j in range(2,sheet.max_column+1):
            dict[sheet.cell(row=1,column=j).value]= sheet.cell(row=i,column=j).value#from excel values trying to build dictionary
            #print(sheet.cell(row=i,column=j).value)#printing all the values from sheet

print(dict)